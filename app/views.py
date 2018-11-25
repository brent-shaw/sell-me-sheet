from flask import render_template, redirect, url_for
from app import app
from openpyxl import load_workbook
import os
from itertools import islice
import datetime

# Global variables
## Site Information
siteName = ""
merchantID = ""
siteID = ""
copyrightYear = 0

## Product catalogue
specials = []
products = []
categories = []

@app.route('/')
@app.route('/load')
def load():
    """Load/reload infgormation from Excel file"""
    #global site info
    global siteName
    global merchantID
    global siteID
    global copyrightYear

    #global catalogue lists
    global specials
    global products
    global categories

    #ensure no duplicates when adding
    specials.clear()
    products.clear()

    #open pricelist file
    wb = load_workbook('pricelist.xlsx')

    #load site information
    ws = wb.get_sheet_by_name('info')
    siteName = ws["B2"].value
    merchantID = ws["B4"].value
    siteID = ws["B5"].value

    #get year
    now = datetime.datetime.now()
    copyrightYear = now.year

    #load specials
    ws = wb.get_sheet_by_name('specials')
    for row in islice(ws.iter_rows(), 1, None):
        specials.append((row[0].value, row[1].value, row[2].value))

    #load products
    ws = wb.get_sheet_by_name('products')
    for row in islice(ws.iter_rows(), 1, None):
        products.append((row[0].value, row[1].value, row[2].value, row[3].value, row[4].value))

    #ensure no duplicates when adding
    categories.clear()

    #load categories
    for i in products:
        cat = i[0]
        if (cat not in categories):
            categories.append(cat)

    return redirect("category/all")

@app.route('/category/<category>')
def index(category):
    """Display items from a specific category.

    By default the site return 'all' items, but allows the catelogue to be filtered by category.

    Keyword arguments:
    category -- which product to return
    """
    global sitename
    global copyrightYear
    return render_template("index.html",
                           name=siteName,
                           year=copyrightYear,
                           specials=specials,
                           products=products,
                           categories=categories,
                           category=category)

@app.route('/item/<c>')
def item(c):
    """Display information about a specific item.

    Keyword arguments:
    c -- the index of the item to be displayed
    """
    global siteName
    global merchantID
    global siteID
    global copyrightYear

    code = int(c)-1
    product = products[code][1]
    price = products[code][2]
    picture = products[code][3]
    description = products[code][4]

    return render_template("item.html",
                           name=siteName,
                           year=copyrightYear,
                           categories=categories,
                           merchantID=merchantID,
                           siteID=siteID,
                           product=product,
                           price=int(price),
                           picture=picture,
                           description=description)
